import pandas as pd
import numpy as np
from scipy.stats import norm
import yfinance as yf
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook

def bs_formula(S,K,r,q, T, sigma):
    d1 = ((np.log(S/K)) + (r-q + 0.5 * sigma ** 2) * T)/(np.sqrt(T)*sigma)
    d2 = d1 - sigma*np.sqrt(T)
    
    call = S * np.exp(-q*T)*norm.cdf(d1) - K*np.exp(-r*T)*norm.cdf(d2)
    put = -S * np.exp(-q*T)*norm.cdf(-d1) + K*np.exp(-r*T)*norm.cdf(-d2)

    return call, put

def vol_implicite(S,K,r,q,T,sigma_implicite, prix_marche_call, prix_marche_put):
    sigma = sigma_implicite
    prix_marche = [prix_marche_call, prix_marche_put]
    vol_imp = [0,0]
    for j in range(2):
        for i in range(1000):
            prix_theorique = bs_formula(S,K,r,q,T,sigma)[j]

            diff = prix_theorique - prix_marche[j]

            if abs(diff) < 1e-10:
                vol_imp[j] = sigma
                
            else:
                d1 = (np.log(S/K) + (r - q + 0.5 * sigma ** 2) * T) / (sigma * np.sqrt(T))
                vega = S * norm.pdf(d1) * np.sqrt(T)
                if vega < 1e-6:
                    break

                sigma = sigma - diff/vega
            
    return vol_imp


def get_data(ticker):

    ticker = yf.Ticker(ticker)
    exp_dates = ticker.options

    all_calls = []
    all_puts = []

    for date in exp_dates:

        option_chain = ticker.option_chain(date)

        call = option_chain.calls
        put = option_chain.puts

        call['expiration'] = date
        put['expiration'] = date

        all_calls.append(call)
        all_puts.append(put)

    all_calls_df = pd.concat(all_calls, ignore_index=True)
    all_puts_df = pd.concat(all_puts, ignore_index=True)

    return all_calls_df, all_puts_df

def time_to_maturity(date):
    maturity_date = datetime.strptime(date, '%Y-%m-%d')
    current_date = datetime.now()
    return (maturity_date-current_date).days/365

def get_filtered_data(liquidity_requirement=250):

    data = get_data("AAPL")
    data_filtered = []
    data_filtered_1 = pd.DataFrame()
    data_filtered_2 = pd.DataFrame()

    for j in range(2):
        
        data_filtered = data[j][data[j]['volume'] > liquidity_requirement]
        
        if j == 0:
            data_filtered_1 = data_filtered
        elif j == 1:
            data_filtered_2 = data_filtered
    
    data_filtered_1_v2 = data_filtered_1.drop(['lastTradeDate','contractSymbol', 'bid', 'ask','openInterest', 'change', 'percentChange', 'volume', 'contractSize','inTheMoney','currency'], axis=1)
    data_filtered_2_v2 = data_filtered_2.drop(['lastTradeDate','contractSymbol', 'bid','ask','openInterest', 'change', 'percentChange', 'volume', 'contractSize','inTheMoney','currency'], axis=1)

    ticker = yf.Ticker("AAPL")
    spot_price = ticker.history(period='1d')['Close'].iloc[-1]
    
    data_filtered_1_v2['SpotPrice'] = spot_price
    data_filtered_2_v2['SpotPrice'] = spot_price

    data_filtered_1_v2 = data_filtered_1_v2.dropna(subset=['strike', 'lastPrice', 'expiration'])
    data_filtered_2_v2 = data_filtered_2_v2.dropna(subset=['strike', 'lastPrice', 'expiration'])

    return data_filtered_1_v2, data_filtered_2_v2

def volatility_smile(data, maturity):
    
    data_maturity = data[data['expiration']==maturity].sort_values(by='strike')

    plt.figure(figsize=(10,6))
    plt.plot(data_maturity['strike'], data_maturity['impliedVolatility'], marker='o', linestyle='-')
    plt.title(f'Smile de Vol pour la maturité {maturity}')
    plt.xlabel('Strike')
    plt.ylabel('Implied Vol')
    plt.show()

def volatility_term_structure(data, strike):

    data_strike = data[data['strike']==strike].sort_values(by='expiration')

    plt.figure(figsize=(10,6))
    plt.plot(data_strike['expiration'], data_strike['impliedVolatility'], marker='o', linestyle='-')
    plt.title(f'Term Structure selon la maturité pour le strike {strike}')
    plt.xlabel('Maturité')
    plt.ylabel('Implied Vol')
    plt.show()

def add_theoritical_value():

    data_calls, data_puts = get_filtered_data(300)

    def compute_theoritical_vol(row, is_call=True):
        S = row['SpotPrice']
        K = row['strike']
        r = 0.05
        q = 0
        T = time_to_maturity(row['expiration'])
        market_price = row['lastPrice']

        vol_imp = vol_implicite(S,K,r,q,T,0.2,market_price, market_price)[0 if is_call else 1]
        return vol_imp
    
    def compute_theoritical_price(row, is_call=True):
        S = row['SpotPrice']
        K = row['strike']
        r = 0.05
        q = 0
        T = time_to_maturity(row['expiration'])
        sigma = row['impliedVolatility']
        th_price = bs_formula(S,K,r,q,T,sigma)[0 if is_call else 1]
        return th_price
    
    data_calls['TheoriticalImpliedVol'] = data_calls.apply(compute_theoritical_vol, axis=1)
    data_puts['TheoriticalImpliedVol'] = data_puts.apply(lambda row:compute_theoritical_vol(row, is_call=False), axis=1)

    data_calls['comparaison'] = data_calls['impliedVolatility']-data_calls['TheoriticalImpliedVol']
    data_puts['comparaison']= data_puts['impliedVolatility']-data_puts['TheoriticalImpliedVol']

    data_calls['TheoriticalPrice'] = data_calls.apply(lambda row:compute_theoritical_price(row, is_call=True), axis=1)
    data_puts['TheoriticalPrice'] = data_puts.apply(lambda row:compute_theoritical_price(row, is_call=False), axis = 1)

    data_calls['comparaisonPrice'] = data_calls['lastPrice']-data_calls['TheoriticalPrice']
    data_puts['comparaisonPrice'] = data_puts['lastPrice']-data_puts['TheoriticalPrice']

    return data_calls, data_puts

def main():

    filename = 'Option-data.xlsx'

    new_order = ['expiration','SpotPrice','strike','impliedVolatility','TheoriticalImpliedVol','comparaison','lastPrice','TheoriticalPrice','comparaisonPrice']

    data_calls = add_theoritical_value()[0]
    data_puts = add_theoritical_value()[1]

    data_calls = data_calls[new_order]
    data_puts = data_puts[new_order]


    with pd.ExcelWriter(filename) as writer:
        data_calls.to_excel(writer, sheet_name='Calls',index=False)
        data_puts.to_excel(writer, sheet_name='Puts', index=False)

    wb = load_workbook(filename)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value))>max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
    wb.save(filename)
    print("Mise en forme de l'excel terminée")

main()